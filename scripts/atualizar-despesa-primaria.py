import json
import tempfile
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[1]

ARQUIVO_SAIDA = (
    BASE / "data" / "economia" / "gerado" /
    "despesa-primaria.json"
)

URL_FONTE = (
    "https://www.tesourotransparente.gov.br/ckan/"
    "dataset/ab56485b-9c40-4efb-8563-9ce3e1973c4b/"
    "resource/527ccdb1-3059-42f3-bf23-b5e3ab4c6dc6/"
    "download/seriehistoricamai26.xlsx"
)

ABA = "2.5-A"
LINHA_ANOS = 5
LINHA_DESPESA = 21
ANO_INICIAL = 2019


def baixar(destino):
    req = urllib.request.Request(
        URL_FONTE,
        headers={
            "User-Agent": "ForaDaPauta/1.0",
            "Accept": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        },
    )

    with urllib.request.urlopen(req, timeout=60) as response:
        conteudo = response.read()

    if len(conteudo) < 10000:
        raise RuntimeError(
            "Arquivo baixado parece pequeno demais."
        )

    destino.write_bytes(conteudo)


def carregar_anterior():
    if not ARQUIVO_SAIDA.exists():
        return {}

    with ARQUIVO_SAIDA.open("r", encoding="utf-8") as f:
        dados = json.load(f)

    return {
        item["ano"]: item["valor"]
        for item in dados.get("anos", [])
        if item.get("valor") is not None
    }


def media(valores):
    if not valores:
        return None

    return round(sum(valores) / len(valores), 2)


print("Atualizando despesa primária...")
print()

anterior = carregar_anterior()

with tempfile.TemporaryDirectory() as tmpdir:
    planilha = Path(tmpdir) / "rtn-serie-historica.xlsx"

    print("Baixando série histórica oficial do Tesouro Nacional...")
    baixar(planilha)

    wb = openpyxl.load_workbook(
        planilha,
        data_only=True,
        read_only=True
    )

    if ABA not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            f"Aba {ABA} não encontrada."
        )

    ws = wb[ABA]

    cabecalho = list(
        next(
            ws.iter_rows(
                min_row=LINHA_ANOS,
                max_row=LINHA_ANOS,
                values_only=True
            )
        )
    )

    linha_despesa = list(
        next(
            ws.iter_rows(
                min_row=LINHA_DESPESA,
                max_row=LINHA_DESPESA,
                values_only=True
            )
        )
    )

    if str(linha_despesa[0]).strip() != "2. DESPESA TOTAL":
        wb.close()
        raise RuntimeError(
            "Linha esperada de DESPESA TOTAL não encontrada."
        )

    serie = {}

    for coluna, ano in enumerate(cabecalho):
        if not isinstance(ano, (int, float)):
            continue

        ano = int(ano)

        if ano < ANO_INICIAL:
            continue

        valor = linha_despesa[coluna]

        if valor is None:
            continue

        serie[ano] = round(float(valor) * 100, 2)

    wb.close()


if not serie:
    raise RuntimeError(
        "Nenhum dado de despesa primária encontrado."
    )


print()
print("Comparando com a versão anterior...")

revisoes = []

for ano, novo in sorted(serie.items()):
    antigo = anterior.get(ano)

    if (
        antigo is not None
        and abs(antigo - novo) > 0.000001
    ):
        revisoes.append({
            "ano": ano,
            "valorAnterior": antigo,
            "valorNovo": novo,
        })

        print(
            f"REVISAO: {ano}: "
            f"{antigo} -> {novo}"
        )

if not revisoes:
    print("Nenhuma revisão histórica detectada.")


anos = []

for ano in sorted(serie):
    registro = {
        "ano": ano,
        "governo": (
            "bolsonaro"
            if ano <= 2022
            else "lula"
        ),
        "valor": serie[ano],
        "tipo": "anual-fechado",
        "origem": (
            "Tesouro Nacional - Resultado do Tesouro Nacional, "
            "Série Histórica, tabela 2.5-A"
        ),
    }

    if ano == 2020:
        registro["contexto"] = "Pandemia de COVID-19"

    anos.append(registro)


bolsonaro_3 = [
    serie[ano]
    for ano in (2019, 2020, 2021)
    if ano in serie
]

lula_3 = [
    serie[ano]
    for ano in (2023, 2024, 2025)
    if ano in serie
]

if len(bolsonaro_3) != 3:
    raise RuntimeError(
        "Comparação Bolsonaro incompleta."
    )

if len(lula_3) != 3:
    raise RuntimeError(
        "Comparação Lula incompleta."
    )


ultimo_ano = max(serie)


dados = {
    "id": "despesa-primaria",
    "titulo": "Despesa primária do Governo Central",
    "unidade": "% do PIB",
    "metodologia": (
        "Despesa total primária do Governo Central apurada "
        "pelo critério de valor pago, como proporção do PIB, "
        "conforme a série histórica do Resultado do Tesouro Nacional."
    ),
    "fonte": {
        "instituicao": "Tesouro Nacional",
        "pesquisa": "Resultado do Tesouro Nacional - Série Histórica",
        "tabela": "2.5-A",
        "url": URL_FONTE,
    },
    "anos": anos,
    "comparacaoMesmaDuracao": {
        "descricao": (
            "Média da despesa primária nos primeiros três "
            "anos completos de cada governo"
        ),
        "bolsonaro": {
            "periodo": "2019-2021",
            "media": media(bolsonaro_3),
        },
        "lula": {
            "periodo": "2023-2025",
            "media": media(lula_3),
        },
    },
    "ultimoAnoDisponivel": ultimo_ano,
    "revisoesDetectadas": revisoes,
    "atualizadoEm": datetime.now().isoformat(
        timespec="seconds"
    ),
}


tmp = ARQUIVO_SAIDA.with_suffix(".json.tmp")

with tmp.open("w", encoding="utf-8") as f:
    json.dump(
        dados,
        f,
        ensure_ascii=False,
        indent=2
    )

with tmp.open("r", encoding="utf-8") as f:
    teste = json.load(f)

if not teste["anos"]:
    raise RuntimeError(
        "Validação final encontrou série vazia."
    )

tmp.replace(ARQUIVO_SAIDA)


print()
print("Despesa primária atualizada com sucesso.")
print()

for ano, valor in sorted(serie.items()):
    print(ano, valor)

print()
print(
    "Bolsonaro 2019-2021 média:",
    dados["comparacaoMesmaDuracao"]
    ["bolsonaro"]["media"]
)

print(
    "Lula 2023-2025 média:",
    dados["comparacaoMesmaDuracao"]
    ["lula"]["media"]
)

print()
print("Arquivo:", ARQUIVO_SAIDA)
