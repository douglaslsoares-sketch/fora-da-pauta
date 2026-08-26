import json
import tempfile
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[1]

ARQUIVO_SAIDA = (
    BASE / "data" / "economia" / "gerado" /
    "carga-tributaria.json"
)

URL_FONTE = (
    "https://www.gov.br/receitafederal/pt-br/"
    "centrais-de-conteudo/publicacoes/estudos/"
    "carga-tributaria/"
    "tabelas-carga-tributaria-no-brasil-2024/"
    "@@download/file"
)

ABA = "T01B"
LINHA_ANOS = 5
LINHA_TOTAL = 6
ANO_INICIAL = 2019


def baixar(destino):
    req = urllib.request.Request(
        URL_FONTE,
        headers={
            "User-Agent": "ForaDaPauta/1.0",
            "Accept": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        },
    )

    with urllib.request.urlopen(req, timeout=60) as response:
        conteudo = response.read()

    if len(conteudo) < 10000:
        raise RuntimeError(
            "Arquivo baixado parece pequeno demais."
        )

    destino.write_bytes(conteudo)


def carregar_anterior():
    if not ARQUIVO_SAIDA.exists():
        return {}

    with ARQUIVO_SAIDA.open("r", encoding="utf-8") as f:
        dados = json.load(f)

    return {
        item["ano"]: item["valor"]
        for item in dados.get("anos", [])
        if item.get("valor") is not None
    }


def media(valores):
    if not valores:
        return None

    return round(sum(valores) / len(valores), 2)


print("Atualizando carga tributária...")
print()

anterior = carregar_anterior()

with tempfile.TemporaryDirectory() as tmpdir:
    planilha = Path(tmpdir) / "carga-tributaria.xlsx"

    print("Baixando planilha oficial da Receita Federal...")
    baixar(planilha)

    wb = openpyxl.load_workbook(
        planilha,
        data_only=True,
        read_only=True
    )

    if ABA not in wb.sheetnames:
        raise RuntimeError(
            f"Aba {ABA} não encontrada."
        )

    ws = wb[ABA]

    cabecalho = list(
        next(
            ws.iter_rows(
                min_row=LINHA_ANOS,
                max_row=LINHA_ANOS,
                values_only=True
            )
        )
    )

    valores = list(
        next(
            ws.iter_rows(
                min_row=LINHA_TOTAL,
                max_row=LINHA_TOTAL,
                values_only=True
            )
        )
    )

    if str(valores[2]).strip() != "Total da Receita Tributária":
        raise RuntimeError(
            "Linha esperada de Total da Receita Tributária não encontrada."
        )

    serie = {}

    for coluna, ano in enumerate(cabecalho):
        if not isinstance(ano, (int, float)):
            continue

        ano = int(ano)

        if ano < ANO_INICIAL:
            continue

        valor = valores[coluna]

        if valor is None:
            continue

        serie[ano] = round(float(valor) * 100, 2)
    wb.close()

if not serie:
    raise RuntimeError(
        "Nenhum dado de carga tributária encontrado."
    )


#
# REVISÕES
#

print()
print("Comparando com a versão anterior...")

revisoes = []

for ano, novo in sorted(serie.items()):
    antigo = anterior.get(ano)

    if (
        antigo is not None
        and abs(antigo - novo) > 0.000001
    ):
        revisoes.append({
            "ano": ano,
            "valorAnterior": antigo,
            "valorNovo": novo,
        })

        print(
            f"REVISAO: {ano}: "
            f"{antigo} -> {novo}"
        )

if not revisoes:
    print("Nenhuma revisão histórica detectada.")


#
# SÉRIE
#

anos = []

for ano in sorted(serie):
    registro = {
        "ano": ano,
        "governo": (
            "bolsonaro"
            if ano <= 2022
            else "lula"
        ),
        "valor": serie[ano],
        "tipo": "anual-fechado",
        "origem": (
            "Receita Federal - Carga Tributária no Brasil, "
            "Tabela TRIB 01-B"
        ),
    }

    if ano == 2020:
        registro["contexto"] = "Pandemia de COVID-19"

    anos.append(registro)


bolsonaro_3 = [
    serie[ano]
    for ano in (2019, 2020, 2021)
    if ano in serie
]

lula_disponivel = [
    serie[ano]
    for ano in sorted(serie)
    if ano >= 2023
]


if len(bolsonaro_3) != 3:
    raise RuntimeError(
        "Comparação Bolsonaro incompleta."
    )

if not lula_disponivel:
    raise RuntimeError(
        "Nenhum dado disponível para o governo Lula."
    )


ultimo_ano = max(serie)


dados = {
    "id": "carga-tributaria",
    "titulo": "Carga tributária",
    "unidade": "% do PIB",
    "metodologia": (
        "Total da receita tributária como proporção do PIB, "
        "conforme a série histórica oficial publicada pela "
        "Receita Federal."
    ),
    "fonte": {
        "instituicao": "Receita Federal",
        "pesquisa": "Carga Tributária no Brasil 2024",
        "tabela": "TRIB 01-B",
        "arquivo": "Tabelas da publicação",
        "url": URL_FONTE,
    },
    "anos": anos,
    "comparacao": {
        "bolsonaro2019a2021": {
            "periodo": "2019-2021",
            "media": media(bolsonaro_3),
        },
        "lulaDisponivel": {
            "periodo": f"2023-{ultimo_ano}",
            "media": media(lula_disponivel),
        },
    },
    "ultimoAnoDisponivel": ultimo_ano,
    "revisoesDetectadas": revisoes,
    "atualizadoEm": datetime.now().isoformat(
        timespec="seconds"
    ),
}


tmp = ARQUIVO_SAIDA.with_suffix(".json.tmp")

with tmp.open("w", encoding="utf-8") as f:
    json.dump(
        dados,
        f,
        ensure_ascii=False,
        indent=2
    )

with tmp.open("r", encoding="utf-8") as f:
    teste = json.load(f)

if not teste["anos"]:
    raise RuntimeError(
        "Validação final encontrou série vazia."
    )

tmp.replace(ARQUIVO_SAIDA)


print()
print("Carga tributária atualizada com sucesso.")
print()

for ano, valor in sorted(serie.items()):
    print(ano, valor)

print()
print(
    "Bolsonaro 2019-2021 média:",
    dados["comparacao"]
    ["bolsonaro2019a2021"]["media"]
)

print(
    f"Lula 2023-{ultimo_ano} média:",
    dados["comparacao"]
    ["lulaDisponivel"]["media"]
)

print()
print("Arquivo:", ARQUIVO_SAIDA)
