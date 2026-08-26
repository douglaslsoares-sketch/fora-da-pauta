import json
import tempfile
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd

BASE = Path(__file__).resolve().parents[1]

SAIDA = (
    BASE / "data" / "economia" / "gerado" /
    "mortalidade-infantil.json"
)

ANO_INICIAL = 2019
ANO_CORRENTE = datetime.now().year


def url_ano(ano, ext):
    return (
        "https://ftp.ibge.gov.br/"
        "Tabuas_Completas_de_Mortalidade/"
        f"Tabuas_Completas_de_Mortalidade_{ano}/"
        f"{ext}/ambos_os_sexos.{ext}"
    )


def baixar(url, destino):
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "ForaDaPauta/1.0"}
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            destino.write_bytes(response.read())
        return True
    except Exception:
        return False


def ler_xls(arquivo):
    wb = xlrd.open_workbook(arquivo)
    ws = wb.sheet_by_index(0)

    for i in range(ws.nrows):
        if ws.cell_value(i, 0) == 0:
            return float(ws.cell_value(i, 1))

    raise RuntimeError("Idade 0 não encontrada no XLS.")


def ler_xlsx(arquivo):
    wb = openpyxl.load_workbook(
        arquivo,
        data_only=True,
        read_only=True
    )

    try:
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(values_only=True):
            if row and row[0] == 0:
                return float(row[1])

    finally:
        wb.close()

    raise RuntimeError("Idade 0 não encontrada no XLSX.")


def carregar_anterior():
    if not SAIDA.exists():
        return {}

    with SAIDA.open("r", encoding="utf-8") as f:
        dados = json.load(f)

    return {
        item["ano"]: item["valor"]
        for item in dados.get("anos", [])
        if item.get("valor") is not None
    }


print("Atualizando mortalidade infantil...")
print()

anterior = carregar_anterior()
serie = {}
fontes_por_ano = {}


with tempfile.TemporaryDirectory() as tmpdir:
    tmpdir = Path(tmpdir)

    for ano in range(ANO_INICIAL, ANO_CORRENTE + 1):
        encontrado = False

        for ext in ("xlsx", "xls"):
            url = url_ano(ano, ext)
            destino = tmpdir / f"mortalidade-{ano}.{ext}"

            if not baixar(url, destino):
                continue

            if ext == "xlsx":
                valor = ler_xlsx(destino)
            else:
                valor = ler_xls(destino)

            serie[ano] = round(valor, 2)
            fontes_por_ano[ano] = {
                "formato": ext,
                "url": url,
            }

            print(
                f"{ano}: encontrado ({ext}) -> "
                f"{serie[ano]}"
            )

            encontrado = True
            break

        if not encontrado:
            print(f"{ano}: ainda não publicado")


if not serie:
    raise RuntimeError(
        "Nenhuma tábua de mortalidade foi encontrada."
    )


#
# REVISÕES
#

print()
print("Comparando com a versão anterior...")

revisoes = []

for ano, novo in sorted(serie.items()):
    antigo = anterior.get(ano)

    if (
        antigo is not None
        and abs(antigo - novo) > 0.000001
    ):
        revisoes.append({
            "ano": ano,
            "valorAnterior": antigo,
            "valorNovo": novo,
        })

        print(
            f"REVISAO: {ano}: "
            f"{antigo} -> {novo}"
        )

if not revisoes:
    print("Nenhuma revisão histórica detectada.")


#
# SÉRIE
#

anos = []

for ano in sorted(serie):
    registro = {
        "ano": ano,
        "governo": (
            "bolsonaro"
            if ano <= 2022
            else "lula"
        ),
        "valor": serie[ano],
        "tipo": "anual-fechado",
        "origem": (
            "IBGE - Tábuas Completas de Mortalidade, "
            "ambos os sexos"
        ),
    }

    if ano == 2020:
        registro["contexto"] = "Pandemia de COVID-19"

    anos.append(registro)


ultimo_ano = max(serie)

anos_sem_dado = [
    {
        "ano": ano,
        "motivo": (
            "Tábua Completa de Mortalidade ainda não "
            "publicada pelo IBGE."
        ),
    }
    for ano in range(ANO_INICIAL, ANO_CORRENTE + 1)
    if ano not in serie
]


if 2019 not in serie or 2022 not in serie:
    raise RuntimeError(
        "Série Bolsonaro incompleta."
    )

variacao_bolsonaro = round(
    serie[2022] - serie[2019],
    2
)


anos_lula = [
    ano
    for ano in serie
    if ano >= 2023
]

if not anos_lula:
    raise RuntimeError(
        "Nenhum dado disponível para o governo Lula."
    )

ultimo_lula = max(anos_lula)

variacao_lula = round(
    serie[ultimo_lula] - serie[2023],
    2
)


dados = {
    "id": "mortalidade-infantil",
    "titulo": "Mortalidade infantil",
    "unidade": "por mil",
    "metodologia": (
        "Probabilidade de morte entre o nascimento e "
        "a idade exata de 1 ano, por mil, segundo as "
        "Tábuas Completas de Mortalidade do IBGE, "
        "para ambos os sexos."
    ),
    "fonte": {
        "instituicao": "IBGE",
        "pesquisa": "Tábuas Completas de Mortalidade",
        "sexo": "Ambos os sexos",
        "arquivosPorAno": fontes_por_ano,
    },
    "anos": anos,
    "anosSemDado": anos_sem_dado,
    "evolucaoDadosDisponiveis": {
        "bolsonaro": {
            "periodo": "2019–2022",
            "inicio": serie[2019],
            "fim": serie[2022],
            "variacaoPorMil": variacao_bolsonaro,
        },
        "lula": {
            "periodo": f"2023–{ultimo_lula}",
            "inicio": serie[2023],
            "fim": serie[ultimo_lula],
            "variacaoPorMil": variacao_lula,
        },
    },
    "ultimoAnoDisponivel": ultimo_ano,
    "revisoesDetectadas": revisoes,
    "atualizadoEm": datetime.now().isoformat(
        timespec="seconds"
    ),
}


tmp = SAIDA.with_suffix(".json.tmp")

with tmp.open("w", encoding="utf-8") as f:
    json.dump(
        dados,
        f,
        ensure_ascii=False,
        indent=2
    )

with tmp.open("r", encoding="utf-8") as f:
    teste = json.load(f)

if not teste["anos"]:
    raise RuntimeError(
        "Validação final encontrou série vazia."
    )

tmp.replace(SAIDA)


print()
print("Mortalidade infantil atualizada com sucesso.")
print()

for ano in range(ANO_INICIAL, ANO_CORRENTE + 1):
    if ano in serie:
        print(ano, serie[ano])
    else:
        print(ano, "sem dado")

print()
print(
    "Bolsonaro 2019-2022:",
    serie[2019],
    "->",
    serie[2022],
    "=",
    variacao_bolsonaro,
    "por mil"
)

print(
    f"Lula 2023-{ultimo_lula}:",
    serie[2023],
    "->",
    serie[ultimo_lula],
    "=",
    variacao_lula,
    "por mil"
)

print()
print("Arquivo:", SAIDA)
